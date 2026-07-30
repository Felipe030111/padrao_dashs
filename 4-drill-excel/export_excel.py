# -*- coding: utf-8 -*-
"""
Exportação em Excel (.xlsx) com openpyxl — usado pelo drill.

    from export_excel import xlsx_response
    @app.route("/export/<obra>.xlsx")
    @auth.login_required
    def export(obra):
        linhas = buscar_detalhe(obra)          # [[c1,c2,...], ...]
        return xlsx_response(["Conta","Realizado","Comp."], linhas,
                             filename=f"detalhe_{obra}.xlsx", sheet="Detalhe")
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from flask import Response


def rows_to_xlsx(headers, rows, sheet="Dados"):
    """Gera um .xlsx (bytes) com cabeçalho em negrito e colunas auto-ajustadas."""
    wb = Workbook(); ws = wb.active; ws.title = sheet[:31]
    head_fill = PatternFill("solid", fgColor="005330")
    head_font = Font(bold=True, color="FFFFFF")
    ws.append(list(headers))
    for c in ws[1]:
        c.font = head_font; c.fill = head_fill
        c.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append(list(r))
    # largura das colunas pelo maior conteúdo
    for i, _ in enumerate(headers, 1):
        col = ws.column_dimensions[ws.cell(row=1, column=i).column_letter]
        longest = max([len(str(headers[i-1]))] + [len(str(r[i-1])) for r in rows if i-1 < len(r)] + [8])
        col.width = min(longest + 3, 60)
    ws.freeze_panes = "A2"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()


def xlsx_response(headers, rows, filename="dados.xlsx", sheet="Dados"):
    """Response Flask que faz o navegador baixar o .xlsx."""
    data = rows_to_xlsx(headers, rows, sheet)
    return Response(
        data,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
